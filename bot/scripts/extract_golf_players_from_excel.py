#!/usr/bin/env python3
"""
Extract embedded photos for golf players from Excel file cells
"""

import json
import logging
import os
import sys
import re
from datetime import datetime
from PIL import Image
import io
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage

# Add the bot directory to the path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Base directory for saving photos
BASE_PHOTO_DIR = "static/logos/players/golf"

# Quality filters
MIN_WIDTH = 1  # Lowered from 200
MIN_HEIGHT = 1  # Lowered from 200
MAX_WIDTH = 2000
MAX_HEIGHT = 2000
MIN_ASPECT_RATIO = 0.5
MAX_ASPECT_RATIO = 2.0


def is_valid_person_image(image_data: bytes) -> tuple[bool, dict]:
    """Validate if an image is a good quality person photo."""
    try:
        # Open image from bytes
        image = Image.open(io.BytesIO(image_data))

        # Get image info
        width, height = image.size
        format_type = image.format
        mode = image.mode

        # Basic size validation (removed restriction)
        # if width < MIN_WIDTH or height < MIN_HEIGHT:
        #     return False, {"error": "Image too small", "width": width, "height": height}

        if width > MAX_WIDTH or height > MAX_HEIGHT:
            return False, {"error": "Image too large", "width": width, "height": height}

        # Aspect ratio validation
        aspect_ratio = width / height
        if aspect_ratio < MIN_ASPECT_RATIO or aspect_ratio > MAX_ASPECT_RATIO:
            return False, {
                "error": "Aspect ratio out of range",
                "aspect_ratio": aspect_ratio,
            }

        # Format validation
        if format_type not in ["JPEG", "PNG", "WEBP"]:
            return False, {"error": "Unsupported format", "format": format_type}

        # Mode validation (should be RGB or RGBA)
        if mode not in ["RGB", "RGBA"]:
            return False, {"error": "Unsupported color mode", "mode": mode}

        # Create metadata
        metadata = {
            "width": width,
            "height": height,
            "format": format_type,
            "mode": mode,
            "aspect_ratio": round(aspect_ratio, 2),
            "file_size": len(image_data),
            "timestamp": datetime.now().isoformat(),
        }

        return True, metadata

    except Exception as e:
        return False, {"error": str(e)}


def save_image(image_data: bytes, file_path: str, metadata: dict) -> bool:
    """Save an image to the specified path."""
    try:
        # Ensure directory exists
        os.makedirs(os.path.dirname(file_path), exist_ok=True)

        # Save image
        with open(file_path, "wb") as f:
            f.write(image_data)

        # Save metadata
        metadata_file = file_path.replace(".png", "_metadata.json")
        with open(metadata_file, "w") as f:
            json.dump(metadata, f, indent=2, default=str)

        return True
    except Exception as e:
        logger.error(f"Error saving image: {e}")
        return False


def clean_filename(name: str) -> str:
    """Clean a player name for use as a filename."""
    cleaned = re.sub(r'[<>:"/\\|?*]', "", name)
    cleaned = cleaned.replace(" ", "_")
    return cleaned


def extract_golf_players_from_excel(excel_file_path: str):
    """Main function to extract embedded photos for golf players from Excel file."""
    logger.info(
        f"Starting golf player photo extraction from Excel file: {excel_file_path}"
    )

    # Load the Excel workbook
    try:
        workbook = openpyxl.load_workbook(excel_file_path)
        logger.info(
            f"Loaded Excel file with {len(workbook.sheetnames)} sheets: {workbook.sheetnames}"
        )

        # Use the first sheet by default
        sheet = workbook.active
        logger.info(f"Using sheet: {sheet.title}")

    except Exception as e:
        logger.error(f"Error reading Excel file: {e}")
        return

    # Create base directory
    os.makedirs(BASE_PHOTO_DIR, exist_ok=True)

    # Track progress
    successful_extractions = 0
    failed_extractions = 0

    # Get all images in the sheet
    images = sheet._images
    logger.info(f"Found {len(images)} embedded images in the sheet")

    # Process each image
    for i, img in enumerate(images):
        try:
            logger.info(f"Processing image {i+1}/{len(images)}")

            # Get image data
            image_data = img._data()

            # Get image position (which cell it's in)
            anchor = img.anchor
            if hasattr(anchor, "_from"):
                cell_position = anchor._from
                row = cell_position.row
                col = cell_position.col
            else:
                # Fallback for different anchor types
                row = getattr(anchor, "row", i + 1)
                col = getattr(anchor, "col", 1)

            logger.info(f"Image {i+1} is in cell {row}, {col}")

            # Try to get player name from nearby cells
            player_name = None
            league = "UNKNOWN"

            # Look for player name in the same row or nearby rows
            for search_row in range(max(1, row - 2), min(sheet.max_row + 1, row + 3)):
                for search_col in range(
                    max(1, col - 2), min(sheet.max_column + 1, col + 3)
                ):
                    cell_value = sheet.cell(row=search_row, column=search_col).value
                    if cell_value and isinstance(cell_value, str):
                        # Check if this looks like a player name
                        if len(cell_value.strip()) > 2 and " " in cell_value:
                            # This might be a player name
                            if not player_name:
                                player_name = cell_value.strip()
                                logger.info(
                                    f"Found potential player name: {player_name}"
                                )
                            # Check if this looks like a league name
                            elif any(
                                league_term in cell_value.upper()
                                for league_term in ["PGA", "LPGA", "TOUR"]
                            ):
                                league = cell_value.strip().upper()
                                logger.info(f"Found potential league: {league}")

            # If no player name found, use a generic name
            if not player_name:
                player_name = f"Player_{i+1}"
                logger.warning(
                    f"No player name found for image {i+1}, using: {player_name}"
                )

            # Validate the image
            is_valid, metadata = is_valid_person_image(image_data)

            if not is_valid:
                logger.warning(
                    f"Image {i+1} failed validation: {metadata.get('error', 'Unknown error')}"
                )
                failed_extractions += 1
                continue

            logger.info(
                f"Valid image for {player_name}: {metadata['width']}x{metadata['height']}"
            )

            # Create file path
            clean_name = clean_filename(player_name)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_path = os.path.join(
                BASE_PHOTO_DIR, league, f"{clean_name}_excel_embedded_{timestamp}.png"
            )

            # Save the image
            if save_image(image_data, file_path, metadata):
                logger.info(f"Successfully extracted photo for {player_name}")
                successful_extractions += 1
            else:
                logger.warning(f"Failed to save photo for {player_name}")
                failed_extractions += 1

        except Exception as e:
            logger.error(f"Error processing image {i+1}: {e}")
            failed_extractions += 1

    logger.info(
        f"Extraction complete! Successful: {successful_extractions}, Failed: {failed_extractions}"
    )


if __name__ == "__main__":
    # You can specify the Excel file path here or pass it as an argument
    excel_file = "golf_players.xlsx"  # Change this to your Excel file path

    if len(sys.argv) > 1:
        excel_file = sys.argv[1]

    extract_golf_players_from_excel(excel_file)
